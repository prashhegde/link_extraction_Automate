# from inputlinkextractor.google import IGoogle
from input_link_extraction.utilities.utils.logger import LogJ
from selenium import webdriver
from input_link_extraction.utilities.xls_generator import *
from time import sleep
from datetime import datetime
from random import randint
from openpyxl import load_workbook

from input_link_extraction.sites.aclens import IAclens

import threading
from selenium import webdriver

TAG = "LinkExtractor"
PATH_INPUT = "C:/Users/Prash/Desktop/Work/scrapj-new/scrapj/data/input/"


class LinkExtractor:
    driver = None
    last_thread = None
    total_links = 0
    total_exp = 0
    start_time = ''
    end_time = ''
    error_logger = LogJ(TAG, "ERROR", True)
    info_logger = LogJ(TAG, "INFO", True)
    websites = []
    wb = load_workbook(filename=PATH_INPUT + 'websites.xlsx', read_only=True)
    ws = wb[wb.get_sheet_names()[0]]
    for row in ws.rows:
        for cell in row:
            if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>"):
                if cell.value is not None and not "":
                    websites.append(cell.value)
                    print(' '.join(websites))
    brands = []
    wb = load_workbook(filename=PATH_INPUT + 'parameters.xlsx', read_only=True)
    ws = wb[wb.get_sheet_names()[0]]
    for row in ws.rows:
        for cell in row:
            if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>"):
                if cell.value is not None and not "":
                    brands.append(cell.value)
                    print(' '.join(brands))
    blacklisted_keys = ['systane', 'eye drop']

    def start_driver(self):
        print('starting driver...')
        try:
            
            driver = webdriver.Chrome("./utilities/drivers/chromedriver.exe")
        except:
            try:
                driver = webdriver.Chrome("./utilities/drivers/chromedriver")
            except:
                print("No Driver")

        sleep(4)
        return driver


    # Open headless chromedriver

# Close chromedriver\
    def close_driver(driver):
        print('closing driver...')
        quit()
        print('closed!')


    def extract_data(self):
        start_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        self.driver = self.start_driver()
        website="www.aclens.com"
        excel = Excel(website, True)
        excel_dump = Excel(website + "_dump", True)
        wb = getattr(excel, "wb")
        excel.add_headers(excel.wb.active,["Product Link", "Product Name", "No. of Reviews"])
        excel_dump.add_headers(excel_dump.wb.active,["Product Link", "Product Name", "No. of Reviews"])
        IAclens(website, self.brands, excel, excel_dump, self.driver, self.match_product_name, self.get_page, self.total_links)
            # if website.__contains__("walgreens"):
            #     IWalgreens(website, self.brands, excel, excel_dump, self.driver, self.logging, self.match_product_name,
            #                self.get_page, self.total_links)
            # if website.__contains__("getlens"):
            #     IGetLens1800(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                  self.match_product_name,
            #                  self.get_page, self.total_links)
            # elif website.__contains__("google"):
            #     IGoogle(website, self.brands, excel, excel_dump, self.driver, self.logging, self.match_product_name,
            #             self.get_page, self.total_links)
            # if website.__contains__("visiondirect"):
            #     IVisionDirect(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                   self.match_product_name,
            #                   self.get_page, self.total_links)
            # if website.__contains__("coastal"):
            #     ICoastal(website + "/contact-lenses", self.brands, excel, excel_dump, self.driver, self.logging,
            #              self.match_product_name,
            #              self.get_page, self.total_links)
            # if website.__contains__("aclens.com"):
            #     IAclens(website, self.brands, excel, excel_dump, self.driver,self.match_product_name,self.get_page, self.total_links)
            # if website.__contains__("www.lens.com"):
            #     ILens(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #           self.match_product_name,
            #           self.get_page, self.total_links)
            # if website.__contains__("walmart"):
            #     IWalmart(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #              self.match_product_name,
            #              self.get_page, self.total_links)
            # if website.__contains__("lensdiscounters.com"):
            #     ILensdiscounters(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                      self.match_product_name,
            #                      self.get_page, self.total_links)
            # if website.__contains__("discountcontactlenses"):
            #     IDiscountContactLens(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                          self.match_product_name,
            #                          self.get_page, self.total_links)
            # if website.__contains__("www.contactlensking.com"):
            #     IContactLensKing(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                      self.match_product_name,
            #                      self.get_page, self.total_links)
            # if website.__contains__("lensdirect"):
            #     ILensDirect(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                 self.match_product_name,
            #                 self.get_page, self.total_links)
            # if website.__contains__("smartbuyglasses"):
            #     ISmartBuyGlasses(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                      self.match_product_name,
            #                      self.get_page, self.total_links)
            # if website.__contains__("framesdirect"):
            #     IFramesDirect(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                   self.match_product_name,
            #                   self.get_page, self.total_links) #//url not working
            # if website.__contains__("clearly"):
            #     IWebContacts(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                  self.match_product_name,
            #                  self.get_page, self.total_links)  #//site redirected
            # if website.__contains__("bestpricecontacts"):
            #     IBestPriceContacts(website, self.brands, excel, excel_dump, self.driver, self.logging,
            #                        self.match_product_name,
            #                        self.get_page, self.total_links)
        threading.Thread(target=self.save_excels, args=(excel, excel_dump)).start()

        print("Input Link Generation Start Time : " + start_time)
        print("Input Link Generation End Time : " + datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
        self.close_driver()

    # Tell the browser to get a page
    def get_page(self, url):
        print('getting page...')
        start_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        print("Start Time" + start_time)
        self.driver.get(url)
        sleep(randint(2, 3))


    def match_product_name(self, product_name):
        product_name = str(product_name).lower()
        wb = load_workbook(filename=PATH_INPUT + 'parameters.xlsx', read_only=True)
        ws = wb["Products"]

        for row in ws.rows:
            for cell in row:
                if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>") and cell.row != 1:
                    if cell.column == 1:
                        name = str(cell.value).lower()
                        if product_name.__contains__(name):
                            for blacklist_key in self.blacklisted_keys:
                                if product_name.__contains__(blacklist_key):
                                    return False
                            return True
        return False

    def save_excels(self, excel, excel_dump):
        while self.last_thread and self.last_thread.isAlive():
            pass
        while True:
            try:
                excel.save_xls(excel.wb)
                excel_dump.save_xls(excel_dump.wb)
                break
            except Exception as e:
                print(e)
                pass


link_extractor = LinkExtractor()
link_extractor.extract_data()
