# from inputlinkextractor.google import IGoogle
from input_link_extraction.utilities.utils.logger import LogJ
from input_link_extraction.utilities.xls_generator import *
from time import sleep
from datetime import datetime
from random import randint
from openpyxl import load_workbook

from input_link_extraction.sites.coastal import ICoastal
import threading
from selenium import webdriver
from PyQt5 import QtCore

TAG = "LinkExtractor"
PATH_INPUT = "../data/input/"


class LinkExtractor:
    driver = None
    last_thread = None
    total_links = 0
    total_exp = 0
    start_time = ''
    end_time = ''
    error_logger = LogJ(TAG, "ERROR", True)
    info_logger = LogJ(TAG, "INFO", True)
    websites = []
    wb = load_workbook(filename=PATH_INPUT + 'websites.xlsx', read_only=True)
    ws = wb[wb.get_sheet_names()[0]]
    for row in ws.rows:
        for cell in row:
            if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>"):
                if cell.value is not None and not "":
                    websites.append(cell.value)
                    print(' '.join(websites))
    brands = []
    wb = load_workbook(filename=PATH_INPUT + 'parameters.xlsx', read_only=True)
    ws = wb[wb.get_sheet_names()[0]]
    for row in ws.rows:
        for cell in row:
            if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>"):
                if cell.value is not None and not "":
                    brands.append(cell.value)
                    print(' '.join(brands))
    blacklisted_keys = ['systane', 'eye drop']

    def start_driver(self):
        print('starting driver...')
        try:
            driver = webdriver.Chrome("./utilities/drivers/chromedriver.exe")
        except:
            try:
                driver = webdriver.Chrome("./utilities/drivers/chromedriver")
            except:
                print("No Driver")

        sleep(4)
        return driver


    # Open headless chromedriver

# Close chromedriver\
    def close_driver(driver):
        print('closing driver...')
        driver.quit()
        print('closed!')


    def extract_data(self):

        start_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        self.driver = self.start_driver()
        website="www.coastal.com"
        excel = Excel(website, True)
        excel_dump = Excel("dumps/" + website + "_dump", True)
        wb = getattr(excel, "wb")
        excel.add_headers(excel.wb.active,["Product Link", "Product Name", "No. of Reviews"])
        excel_dump.add_headers(excel_dump.wb.active,["Product Link", "Product Name", "No. of Reviews"])

        if website.__contains__("coastal"):
            ICoastal(website + "/contact-lenses", self.brands, excel, excel_dump, self.driver, self.logging,
            self.match_product_name,self.get_page, self.total_links)

        threading.Thread(target=self.save_excels, args=(excel, excel_dump)).start()
        print("Input Link Generation Start Time : " + start_time)
        print("Input Link Generation End Time : " + datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
        self.close_driver(self.driver)
        self.finished.emit()

    # Tell the browser to get a page
    def get_page(self, url):
        print('getting page...')
        start_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        print("Start Time" + start_time)
        self.driver.get(url)
        sleep(randint(2, 3))

    def logging(self, curr_links, log, last_thread):
        self.last_thread = last_thread
        while last_thread and last_thread.isAlive():
            pass
        log.append(self.total_links - curr_links)
        self.info_logger.log(log)

    def match_product_name(self, product_name):
        product_name = str(product_name).lower()
        wb = load_workbook(filename=PATH_INPUT + 'parameters.xlsx', read_only=True)
        ws = wb["Products"]

        for row in ws.rows:
            for cell in row:
                if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>") and cell.row != 1:
                    if cell.column == 1:
                        name = str(cell.value).lower()
                        if product_name.__contains__(name):
                            for blacklist_key in self.blacklisted_keys:
                                if product_name.__contains__(blacklist_key):
                                    return False
                            return True
        return False

    def save_excels(self, excel, excel_dump):
        while self.last_thread and self.last_thread.isAlive():
            pass
        while True:
            try:
                excel.save_xls(excel.wb)
                excel_dump.save_xls(excel_dump.wb)
                break
            except Exception as e:
                print(e)
                pass


link_extractor = LinkExtractor()
link_extractor.extract_data()
