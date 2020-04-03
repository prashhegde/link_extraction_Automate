from random import randint
from time import sleep
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
from review_extraction.utilities.utils.logger import LogJ
import threading
from random import randint
from time import sleep
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
from review_extraction.utilities.utils.logger import LogJ
import threading
from selenium import webdriver
from review_extraction.utilities.xls_generator import *
from review_extraction.aclens import *
from pathlib import Path


TAG = "www.aclens.com"
PATH_INPUT = "C:/Users/Prash/Desktop/Work/scrapj-new/scrapj/data/input/"


class AcLens():
    driver = None
    total_reviews = 0
    total_exp = 0
    start_time = ''
    end_time = ''
    params = ()
    error_logger = LogJ(TAG, "ERROR")
    info_logger = LogJ(TAG, "INFO")

    def __init__(self, driver, params):
        self.driver = driver
        self.params = params

    def extract_data(self, excel):
        wb = load_workbook(filename=PATH_INPUT + 'www.aclens.com.xlsx', read_only=True)
        ws = wb[wb.get_sheet_names()[0]]
        excel.add_headers(excel.wb.active,
                          ["Title", "Comments", "Overall", "Comfort", "Vision", "Value for Money", "Author", "Date",
                           "Pros", "Cons", "Original Source",
                           "Reply from Acuvue", "Product Name", "Product Link", "Website"])
        for row in ws.rows:
            for cell in row:
                if str(type(cell)).__eq__("<class 'openpyxl.cell.read_only.ReadOnlyCell'>") and cell.row != 1:
                    if cell.column == 1:
                        print(cell.row)
                        print(cell.value)
                        try:
                            excel.save_xls(excel.wb)
                        except Exception as e:
                            print(e)
                            self.total_exp += 1
                        # url_to_crawl = "http://www.visiondirect.co.uk/brand/acuvue-contact-lenses/1-day-acuvue-moist"
                        # self.get_page(url_to_crawl)
                        if cell.value is not None and not "":
                            self.get_page(cell.value)
                            self.grab_reviews(excel, row[1].value, cell.value)

    # Tell the browser to get a page
    def get_page(self, url):
        print('getting page...')
        start_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        print("Start Time" + start_time)
        sleep(randint(2, 5))
        self.driver.get(url)

    def grab_reviews(self, excel, product_name, product_url):
        print('grabbing reviews......')
        print(datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
        log = [product_name, product_url]
        curr_reviews = self.total_reviews
        last_thread = None

        global prev_div
        prev_div = ''
        try:
            product_name = self.driver.find_element_by_xpath('//div[contains(@class ,"dcl-lens-detail__lens-info-image")]/div/h1')
            product_name = str(product_name.get_attribute("innerHTML")).strip()
        except:
            print("Error getting product name")
            pass
        while True:
            try:
                sort_button = self.driver.find_element_by_xpath('//div[contains(@class, "yotpo-reviews-header") and contains(@class, "active")]/div/div/div/div')
                self.driver.execute_script("arguments[0].click();", sort_button)
                sleep(1)
                sort_button = self.driver.find_element_by_xpath('//div[contains(@class, "yotpo-reviews-header") and contains(@class, "active")]/div/div/div/ul/li[2]/a')
                self.driver.execute_script("arguments[0].click();", sort_button)
                sleep(15 / 10)
                elements = self.driver.find_elements_by_xpath(
                    '//*[contains(@id,"yotpo-reviews")]/div[boolean(number(@data-review-id))]')
            except Exception as e:
                #self.logging(curr_reviews, log, None)
                return
            for div in elements:
                cur_div = div.get_attribute("innerHTML")
                if prev_div.__eq__(cur_div):
                    retry = 15
                    while retry >= 0:
                        elements = self.driver.find_elements_by_xpath(
                            '//*[contains(@id,"yotpo-reviews")]/div[boolean(number(@data-review-id))]')
                        try:
                            cur_div = div.get_attribute("innerHTML")
                        except:
                            pass
                        if prev_div.__eq__(cur_div):
                            sleep(1 / 10)
                            retry -= 1
                        else:
                            break
                        if retry == 0:
                            #threading.Thread(target=self.logging, args=(curr_reviews, log, last_thread)).start()
                            return
                try:
                    last_thread = threading.Thread(target=self.thread_process, args=
                    (div.get_attribute("innerHTML"), product_name, product_url, excel))
                    last_thread.start()
                except:
                    self.total_exp += 1
                    print("Error: unable to start thread")
            try:
                prev_div = elements[0].get_attribute("innerHTML")
            except Exception as e:
                print(e)
                pass
            try:
                next_button = self.driver.find_element_by_xpath(
                    '//*[contains(@id,"yotpo-reviews")]//a[contains(@class,"yotpo_next")]')
                # print(next_button.get_attribute("innerHTML"))
                sleep(1)
                self.driver.execute_script("arguments[0].click();", next_button)
                
            except Exception as e:
                print(e)

    def logging(self, curr_reviews, log, last_thread):
        while last_thread and last_thread.isAlive():
            pass
        log.append(self.total_reviews - curr_reviews)
        self.info_logger.log(log)

    def thread_process(self, div, product_name, product_url, excel):
        # row = self.process_elements(div)
        row = self.process_soup(div)
        if row:
            row.append(product_name)
            row.append(product_url)
            row.append(TAG)
            excel.insert_row(getattr(excel, "wb"), row)

    def process_soup(self, div):
        soup = BeautifulSoup(div, 'lxml')
        attributes = []

        try:
            content = soup.select_one(".content-title")
            if content is not None:
                print(content.getText())
                attributes.append(content.getText())
            else:
                attributes.append("NA1")
        except:
            attributes.append("NA1")
            pass
        try:
            content = soup.select_one(".content-review")
            if content is not None:
                print(content.getText())
                attributes.append(content.getText())
            else:
                attributes.append("NA2")
        except:
            attributes.append("NA2")
            pass
        try:
            content = soup.select(".yotpo-icon-star")
            if content is not None:
                star = str(len(content)) + " out of 5"
                print(star)
                attributes.append(star)
            else:
                return []
                attributes.append("NA3")
        except:
            attributes.append("NA3")
            pass
        attributes.append("NA3")
        attributes.append("NA3")
        attributes.append("NA3")
        try:
            content = soup.select_one(".yotpo-user-name")
            if content is not None:
                author = str(content.getText()).strip()
                print(author)
                attributes.append(author)
            else:
                attributes.append("NA4")
        except:
            attributes.append("NA4")
            pass
        try:
            content = soup.select_one(".yotpo-review-date")
            if content is not None:
                date = str(content.getText()).strip()
                print(date)
                date = datetime.strptime(date, '%m/%d/%y').strftime('%Y/%m/%d')
                attributes.append(date)
            else:
                attributes.append("")
        except:
            attributes.append("")
            pass
        attributes.append("NA6")
        attributes.append("NA7")
        attributes.append(TAG)
        attributes.append("NA9")

        self.total_reviews = self.total_reviews + 1
        print("Total Number of reviews : " + str(self.total_reviews))
        print("Total Number of exp : " + str(self.total_exp))
        return attributes

# Open headless chromedriver
def start_driver():
    print('starting driver...')
    try:
        driver = webdriver.Chrome("C:/Users/Prash/Desktop/Work/scrapj-Thread-Implementation/scrapj/input_link_extraction/utilities/drivers/chromedriver.exe")
    except:
        try:
            driver = webdriver.Chrome("C:/Users/Prash/Desktop/Work/scrapj-Thread-Implementation/scrapj/input_link_extraction/utilities/drivers/chromedriver")
        except:
            print("No Driver")

    sleep(4)
    return driver


# Close chromedriver\
def close_driver(driver):
    print('closing driver...')
    driver.quit()
    print('closed!')

start_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
driver = start_driver()

wg = Excel("www.aclens.com") #correct
wb = getattr(wg, "wb")
lens = AcLens(driver, None)
lens.extract_data(wg)
while True:
    try:
        wg.save_xls(wb)
        Path("../status/" + "www.aclens.com" + ".txt").touch()
        break
    except:
        pass